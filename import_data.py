#!/usr/bin/env python3
"""
Inventory Importer
------------------
Feed it an existing spreadsheet and it adds everything into the Inventory app.

Mapping:
    the FILE      -> a category (named after the file)
    each SHEET    -> a subcategory under it   (Excel files)
    each ROW      -> an item
    each COLUMN   -> a field on that item (custom fields are created as needed)

A plain .csv has no sheets, so it becomes a single category named after the file.

Usage:
    python3 import_data.py warehouse.xlsx
    python3 import_data.py stock.csv
    python3 import_data.py warehouse.xlsx --category Equipment
    python3 import_data.py warehouse.xlsx --dry-run
    python3 import_data.py                       (no args -> opens a file picker)

Options:
    --category NAME   Put everything under NAME instead of the file name.
    --dry-run         Show exactly what would be imported, write nothing.
    --sheet NAME      Import only this one sheet.
"""

import re
import csv
import sys
import uuid
import argparse
from pathlib import Path

import app as inv  # reuses the app's data folder, CSV format and history log

# Internal key used to carry a row's spreadsheet fill colour through the
# mapping step. It never becomes a custom field; it lands in the 'color' column.
COLOR_KEY = "__row_color__"


# ---------------------------------------------------------------------------
# Reading source files
# ---------------------------------------------------------------------------

# Column names we recognise as the item's name / quantity (case-insensitive).
NAME_HINTS = ["name", "item", "item name", "title", "description", "product",
              "bezeichnung", "artikel", "gegenstand"]
QTY_HINTS = ["quantity", "qty", "count", "amount", "stock", "menge", "anzahl"]


def _clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    # "nan"/"null" are artefacts of other export tools. "none" is NOT treated as
    # empty - a user may well have typed it as a real answer.
    return "" if s.lower() in ("nan", "null") else s


def read_csv_sheets(path):
    """A CSV is one sheet; its name is the file stem."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open("r", newline="", encoding=enc) as f:
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except Exception:
                    dialect = csv.excel
                rows = [
                    {(_clean(k) or f"column_{i}"): _clean(v)
                     for i, (k, v) in enumerate(r.items()) if k is not None}
                    for r in csv.DictReader(f, dialect=dialect)
                ]
            return [(path.stem, rows)]
        except UnicodeDecodeError:
            continue
        except Exception as e:
            raise SystemExit(f"Could not read {path.name}: {e}")
    raise SystemExit(f"Could not decode {path.name} with any known encoding.")


def cell_color(cell):
    """
    The cell's background fill as '#rrggbb', or '' if it has none.

    Excel stores fills as direct RGB, an indexed palette entry, or a theme
    reference. The first two are exact; theme colours have no RGB value stored
    in the file, so those are reported as no colour rather than guessed at.
    """
    try:
        fill = cell.fill
        if fill is None or fill.patternType in (None, "none"):
            return ""
        color = fill.fgColor or fill.start_color
        if color is None:
            return ""
        if color.type == "rgb" and isinstance(color.rgb, str):
            rgb = color.rgb
            if len(rgb) == 8:          # AARRGGBB - drop the alpha
                if rgb[:2] == "00":    # fully transparent = no fill
                    return ""
                rgb = rgb[2:]
            if len(rgb) == 6 and re.fullmatch(r"[0-9A-Fa-f]{6}", rgb):
                if rgb.upper() == "FFFFFF":
                    return ""          # plain white is not a highlight
                return "#" + rgb.lower()
        if color.type == "indexed":
            from openpyxl.styles.colors import COLOR_INDEX
            idx = color.indexed
            if isinstance(idx, int) and 0 <= idx < len(COLOR_INDEX):
                rgb = COLOR_INDEX[idx]
                if isinstance(rgb, str) and len(rgb) == 8:
                    rgb = rgb[2:]
                    if rgb.upper() not in ("FFFFFF", "000000"):
                        return "#" + rgb.lower()
    except Exception:
        pass
    return ""


def row_color(cells, header_len):
    """
    One colour for the whole item: the most common fill across the row's cells.
    Covers both a coloured row and a coloured column (that column's cell is the
    only filled one in the row).
    """
    seen = {}
    for c in cells[:header_len]:
        col = cell_color(c)
        if col:
            seen[col] = seen.get(col, 0) + 1
    if not seen:
        return ""
    return max(seen.items(), key=lambda kv: kv[1])[0]


def read_excel_sheets(path):
    """Every worksheet becomes its own (sheet_name, rows) pair."""
    try:
        import openpyxl  # noqa: F401
    except ModuleNotFoundError:
        raise SystemExit(
            "Reading Excel files needs the 'openpyxl' package.\n"
            "Install it with:\n\n    python3 -m pip install openpyxl\n"
        )
    from openpyxl import load_workbook

    # read_only=False so cell fills (colours) are available
    wb = load_workbook(path, data_only=True)
    out = []
    for ws in wb.worksheets:
        grid = list(ws.iter_rows())
        if not grid:
            out.append((ws.title, []))
            continue
        # first non-empty row is the header
        start = 0
        for i, row in enumerate(grid):
            if any(_clean(c.value) for c in row):
                start = i
                break
        header = [_clean(c.value) or f"column_{i+1}"
                  for i, c in enumerate(grid[start])]
        rows = []
        for raw in grid[start + 1:]:
            if not any(_clean(c.value) for c in raw):
                continue  # skip blank separator rows
            row = {header[i]: _clean(c.value)
                   for i, c in enumerate(raw) if i < len(header)}
            color = row_color(list(raw), len(header))
            if color:
                row[COLOR_KEY] = color
            rows.append(row)
        out.append((ws.title, rows))
    wb.close()
    return out


def read_sheets(path):
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xltx"):
        return read_excel_sheets(path)
    if path.suffix.lower() in (".csv", ".tsv", ".txt"):
        return read_csv_sheets(path)
    raise SystemExit(f"Unsupported file type: {path.suffix or '(none)'}. "
                     "Use .csv, .tsv, .xlsx or .xlsm.")


# ---------------------------------------------------------------------------
# Mapping rows -> inventory items
# ---------------------------------------------------------------------------

def pick_column(header, hints):
    lowered = {h.lower().strip(): h for h in header}
    for hint in hints:
        if hint in lowered:
            return lowered[hint]
    for h in header:                       # fall back to a partial match
        for hint in hints:
            if hint in h.lower():
                return h
    return None


def row_to_item(row, name_col, qty_col):
    """Turn one spreadsheet row into an inventory item dict."""
    fields = {}
    for key, val in row.items():
        if not key or key == COLOR_KEY or not _clean(val):
            continue
        if key == name_col or key == qty_col:
            continue
        fields[inv.sanitize_field_name(key)] = _clean(val)

    color = row.get(COLOR_KEY, "")
    if color:
        fields["color"] = color        # a real base column, not a custom field

    name = _clean(row.get(name_col)) if name_col else ""
    if not name:
        # no usable name column -> build one from the first filled value
        name = next((v for k, v in ((k, _clean(x)) for k, x in row.items())
                     if v and k != COLOR_KEY), "")
    qty = _clean(row.get(qty_col)) if qty_col else ""
    return name, (qty or "1"), fields


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------

def import_file(path, base_category=None, only_sheet=None, dry_run=False,
                parent=None, quiet=False):
    """
    Import one spreadsheet. Returns a summary dict.

    base_category replaces the file name as the top category; `parent` instead
    keeps the file name and nests it, giving Parent/FileName[/Sheet].
    """
    path = Path(path).expanduser()
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    sheets = read_sheets(path)
    if only_sheet:
        sheets = [(n, r) for n, r in sheets if n == only_sheet]
        if not sheets:
            raise SystemExit(f"No sheet named '{only_sheet}' in {path.name}.")

    if base_category:
        top = base_category.strip().strip("/")
    else:
        stem = path.stem.strip().replace("/", "-")
        top = f"{parent.strip().strip('/')}/{stem}" if parent else stem
    single_sheet_csv = path.suffix.lower() in (".csv", ".tsv", ".txt")

    inv.ensure_files()
    rows, custom = inv.read_inventory()
    cat_map = inv.read_categories()

    summary = []
    added_total = 0
    ts = inv.now_str()

    for sheet_name, sheet_rows in sheets:
        # a CSV has no real sheets -> everything lands directly in the category
        if single_sheet_csv and not base_category:
            cat_path = top
        else:
            clean_sheet = str(sheet_name).strip().replace("/", "-") or "Sheet"
            cat_path = f"{top}/{clean_sheet}"

        header = [k for k in (sheet_rows[0].keys() if sheet_rows else [])
                  if k != COLOR_KEY]
        name_col = pick_column(header, NAME_HINTS)
        qty_col = pick_column(header, QTY_HINTS)
        colored = sum(1 for r in sheet_rows if r.get(COLOR_KEY))

        cat_map.setdefault(top, [])
        cat_map.setdefault(cat_path, [])

        count = 0
        for r in sheet_rows:
            name, qty, fields = row_to_item(r, name_col, qty_col)
            if not name:
                continue  # completely empty row
            for key in fields:
                if key not in inv.BASE_COLUMNS and key not in custom:
                    custom.append(key)
            item = {c: "" for c in inv.BASE_COLUMNS + custom}
            item.update(fields)
            item["id"] = uuid.uuid4().hex[:12]
            item["name"] = name
            item["quantity"] = qty
            item["category_path"] = cat_path
            item["date_added"] = ts
            item["last_modified"] = ts
            rows.append(item)
            count += 1

        added_total += count
        summary.append((cat_path, count, name_col, qty_col,
                        [h for h in header if h not in (name_col, qty_col)],
                        colored))

    result = {
        "file": path.name,
        "path": str(path),
        "top": top,
        "added": added_total,
        "colored": sum(s[5] for s in summary),
        "categories": [{"path": s[0], "items": s[1], "name_from": s[2],
                        "quantity_from": s[3], "extra_fields": s[4],
                        "colored": s[5]} for s in summary],
        "saved": False,
    }

    # ---- report ----
    if not quiet:
        print()
        print("=" * 66)
        print(f"  {'DRY RUN - nothing written' if dry_run else 'Importing'}: {path.name}")
        print("=" * 66)
        print(f"  Category      : {top}")
        print(f"  Data folder   : {inv.get_data_dir()}")
        print()
        for cat_path, count, name_col, qty_col, extras, colored in summary:
            print(f"  {cat_path}")
            print(f"      items       : {count}")
            print(f"      name from   : {name_col or '(first non-empty column)'}")
            print(f"      quantity    : {qty_col or '(defaults to 1)'}")
            if extras:
                shown = ", ".join(extras[:8]) + (" ..." if len(extras) > 8 else "")
                print(f"      extra fields: {shown}")
            if colored:
                print(f"      colours kept: {colored} row(s)")
            print()
        print(f"  Total items: {added_total}")
        print("=" * 66)

    if dry_run:
        if not quiet:
            print("  Nothing was saved. Re-run without --dry-run to import.\n")
        return result

    if added_total == 0:
        if not quiet:
            print("  No rows with usable data were found - nothing imported.\n")
        return result

    inv.write_categories(cat_map, sync_view=False)
    inv.write_inventory(rows, custom)          # this rebuilds by_category/ too
    inv.log_history("import_file", top,
                    f"{path.name}: {added_total} items into "
                    f"{len(summary)} categor{'y' if len(summary)==1 else 'ies'}")
    if not quiet:
        print(f"  Saved. Open the app to see them under '{top}'.\n")
    result["saved"] = True
    return result


# ---------------------------------------------------------------------------
# Folder import: every spreadsheet in a folder, in one go
# ---------------------------------------------------------------------------

FOLDER_SUFFIXES = {".csv", ".tsv", ".xlsx", ".xlsm"}


def _is_app_generated(p):
    """True for the app's own data files, so a data folder can't re-import itself."""
    if p.name in {inv.INVENTORY_CSV, inv.CATEGORIES_CSV, inv.HISTORY_CSV,
                  inv.DELETED_CSV, inv.ORDER_CSV, "_Overview.csv"}:
        return True
    if p.suffix.lower() in (".csv", ".tsv"):
        try:
            with p.open("r", encoding="utf-8-sig", errors="replace") as f:
                if f.readline().startswith(inv.HIER_MARKER[:30]):
                    return True      # a by_category/ mirror file
        except Exception:
            pass
    return False


def find_importable(folder, recursive=False):
    """Spreadsheets in `folder` worth importing, sorted, app files excluded."""
    folder = Path(folder).expanduser()
    it = folder.rglob("*") if recursive else folder.glob("*")
    found = []
    for p in sorted(it):
        if not p.is_file():
            continue
        if p.suffix.lower() not in FOLDER_SUFFIXES:
            continue
        if p.name.startswith(("~$", ".")):
            continue                      # Excel lock files, hidden files
        if _is_app_generated(p):
            continue
        found.append(p)
    return found


def import_folder(folder, under=None, recursive=False, dry_run=False,
                  mirror_subfolders=True, quiet=False):
    """
    Import every spreadsheet in a folder. Each file becomes its own category
    (named after the file); with `under` they all nest beneath that name. When
    recursing, subfolders become parent categories.
    """
    folder = Path(folder).expanduser()
    if not folder.is_dir():
        raise SystemExit(f"Not a folder: {folder}")

    files = find_importable(folder, recursive)
    if not files:
        return {"folder": str(folder), "files": [], "added": 0, "skipped": 0,
                "message": "No .csv/.tsv/.xlsx files to import were found there."}

    if not quiet:
        print()
        print("=" * 66)
        print(f"  {'DRY RUN - nothing written' if dry_run else 'Folder import'}: {folder}")
        print(f"  {len(files)} file(s) found")
        print("=" * 66)

    results, total, failed = [], 0, []
    for f in files:
        # subfolder path becomes a parent category when recursing
        parent = under or None
        if recursive and mirror_subfolders:
            rel = f.parent.relative_to(folder)
            if str(rel) != ".":
                sub = "/".join(p.replace("/", "-") for p in rel.parts)
                parent = f"{under}/{sub}" if under else sub
        try:
            r = import_file(f, parent=parent, dry_run=dry_run, quiet=True)
            results.append(r)
            total += r["added"]
            if not quiet:
                cats = ", ".join(c["path"] for c in r["categories"]) or "-"
                col = f", {r['colored']} coloured" if r["colored"] else ""
                print(f"  {f.name:<34} {r['added']:>5} item(s){col}")
                print(f"  {'':<34} -> {cats}")
        except SystemExit as e:
            failed.append({"file": f.name, "error": str(e)})
            if not quiet:
                print(f"  {f.name:<34}   SKIPPED: {e}")

    if not quiet:
        print("=" * 66)
        print(f"  Total: {total} item(s) from {len(results)} file(s)"
              + (f", {len(failed)} skipped" if failed else ""))
        if dry_run:
            print("  Nothing was saved. Re-run without --dry-run to import.")
        print()

    if not dry_run and total:
        inv.log_history("import_folder", str(folder),
                        f"{total} items from {len(results)} file(s)")
    return {"folder": str(folder), "files": results, "failed": failed,
            "added": total, "skipped": len(failed)}


def pick_file_dialog():
    """Ask the OS for a file when the script is run with no arguments.

    Uses the app's picker, which supports macOS, Windows and Linux.
    """
    try:
        return inv.pick_file_native()
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser(
        description="Import spreadsheets into the Inventory app "
                    "(file -> category, sheet -> subcategory, row -> item). "
                    "Give it a FOLDER to import every spreadsheet inside it.")
    ap.add_argument("file", nargs="?",
                    help="path to a .csv / .xlsx file, or a folder of them")
    ap.add_argument("--category", help="import under this category name")
    ap.add_argument("--sheet", help="import only this sheet (single file only)")
    ap.add_argument("--under", help="folder import: nest everything under this category")
    ap.add_argument("--recursive", action="store_true",
                    help="folder import: include subfolders, mirroring them as categories")
    ap.add_argument("--dry-run", action="store_true",
                    help="preview the import without saving")
    args = ap.parse_args()

    target = args.file or pick_file_dialog()
    if not target:
        ap.print_help()
        return 1

    if Path(target).expanduser().is_dir():
        import_folder(target, under=args.under or args.category,
                      recursive=args.recursive, dry_run=args.dry_run)
    else:
        import_file(target, base_category=args.category,
                    only_sheet=args.sheet, dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    sys.exit(main())
